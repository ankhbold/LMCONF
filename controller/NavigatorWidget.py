# coding=utf8

__author__ = 'B.Ankhbold'

from qgis.core import *
from PyQt4.QtGui import *
from PyQt4.QtCore import *
from PyQt4.QtXml import *
from PyQt4.QtCore import QDate
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy import func, or_, and_, desc
from sqlalchemy.sql.expression import cast
from sqlalchemy import func
from ..view.Ui_NavigatorWidget import Ui_NavigatorWidget
from datetime import timedelta
from xlsxwriter.utility import xl_rowcol_to_cell, xl_col_to_name
import xlsxwriter
import time
import urllib
import urllib2
import json
import os
from openpyxl import Workbook
import openpyxl

LANDUSE_1 = u'Хөдөө аж ахуйн газар'
LANDUSE_2 = u'Хот, тосгон, бусад суурины газар'
LANDUSE_3 = u'Зам, шугам сүлжээний газар'
LANDUSE_4 = u'Ойн сан бүхий газар'
LANDUSE_5 = u'Усны сан бүхий газар'
LANDUSE_6 = u'Улсын тусгай хэрэгцээний газар'

class NavigatorWidget(QDockWidget, Ui_NavigatorWidget):
    def __init__(self, plugin, parent=None):
        super(NavigatorWidget, self).__init__(parent)

        self.setupUi(self)
        self.setAllowedAreas(Qt.LeftDockWidgetArea | Qt.RightDockWidgetArea)
        self.plugin = plugin
        self.au1_code = None
        self.au2_code = None
        self.au3_code = None
        self.file = None
        self.__au_admin()

    def __au_admin(self):

        self.working_l1_cbox.clear()
        au1_url = urllib.urlopen("http://egazar.gov.mn:8060/api/geo/json/level/one/wihtoutgeom")
        au1_data = json.load(au1_url)
        au1_data_sort = sorted(au1_data['features'], key=lambda i: (i['properties'], i['type']))
        for value in au1_data_sort:
            value_prop = value['properties']
            code = value_prop['code']
            name = value_prop['name']
            self.au1_code = code
            self.working_l1_cbox.addItem(code + ':' + name, code)

        layer_list = []
        layers = QgsMapLayerRegistry.instance().mapLayers()
        for id, layer in layers.iteritems():
            if layer.name() == u"Аймгийн хил":
                layer_list.append(id)

        QgsMapLayerRegistry.instance().removeMapLayers(layer_list)

        self.plugin.iface.addVectorLayer(
            'http://egazar.gov.mn:8060/api/geo/json/level/one',
            u'Аймгийн хил', 'ogr')
        layers = QgsMapLayerRegistry.instance().mapLayers()
        for id, layer in layers.iteritems():
            if layer.name() == u"Аймгийн хил":
                layer.loadNamedStyle(
                    str(os.path.dirname(os.path.realpath(__file__))[:-10]) + "template\style/au_level1.qml")

    @pyqtSlot(int)
    def on_working_l1_cbox_currentIndexChanged(self, index):

        self.au1_code = self.working_l1_cbox.itemData(self.working_l1_cbox.currentIndex())
        self.working_l2_cbox.clear()
        params = urllib.urlencode({'au1_code': self.au1_code})
        au2_url = urllib.urlopen("http://egazar.gov.mn:8060/api/geo/json/level/two/by/level1", params)
        au2_data = json.load(au2_url)
        au2_data_sort = sorted(au2_data['features'], key=lambda i: (i['properties'], i['type']))

        for value in au2_data_sort:
            value_prop = value['properties']
            code = value_prop['code']
            name = value_prop['name']
            self.au2_code = code
            self.working_l2_cbox.addItem(code + ':' + name, code)

        layer_list = []
        layers = QgsMapLayerRegistry.instance().mapLayers()
        for id, layer in layers.iteritems():
            if layer.name() == u"Сумын хил":
                layer_list.append(id)

        QgsMapLayerRegistry.instance().removeMapLayers(layer_list)

        self.plugin.iface.addVectorLayer(
            'http://egazar.gov.mn:8060/api/geo/json/level/two/by/level1?au1_code' + '=' + str(self.au1_code),
            u'Сумын хил', 'ogr')
        layers = QgsMapLayerRegistry.instance().mapLayers()
        for id, layer in layers.iteritems():
            if layer.name() == u"Сумын хил":
                layer.loadNamedStyle(
                    str(os.path.dirname(os.path.realpath(__file__))[:-10]) + "template\style/au_level2.qml")

    @pyqtSlot(int)
    def on_working_l2_cbox_currentIndexChanged(self, index):

        self.au2_code = self.working_l2_cbox.itemData(self.working_l2_cbox.currentIndex())

        layer_list = []
        layers = QgsMapLayerRegistry.instance().mapLayers()
        for id, layer in layers.iteritems():
            if layer.name() == u"Нэгж талбар":
                layer_list.append(id)
            if layer.name() == u"Барилга":
                layer_list.append(id)

        QgsMapLayerRegistry.instance().removeMapLayers(layer_list)

        self.plugin.iface.addVectorLayer(
            'http://egazar.gov.mn:8060/api/geo/json/parcel/by/soum?soum_code' + '=' + str(self.au2_code),
            u'Нэгж талбар', 'ogr')

        self.plugin.iface.addVectorLayer(
            'http://egazar.gov.mn:8060/api/geo/json/building/by/soum?soum_code' + '=' + str(self.au2_code),
            u'Барилга', 'ogr')

        layer_active = None
        layers = QgsMapLayerRegistry.instance().mapLayers()
        for id, layer in layers.iteritems():
            if layer.name() == u"Нэгж талбар":
                layer_active = layer
                layer.loadNamedStyle(
                    str(os.path.dirname(os.path.realpath(__file__))[:-10]) + "template\style/ca_parcel.qml")
            if layer.name() == u"Барилга":
                layer.loadNamedStyle(
                    str(os.path.dirname(os.path.realpath(__file__))[:-10]) + "template\style/ca_building.qml")

        self.plugin.iface.setActiveLayer(layer_active)
        self.plugin.iface.zoomToActiveLayer()

    @pyqtSlot()
    def on_load_exc_button_clicked(self):

        file_dialog = QFileDialog()
        file_dialog.setWindowFlags(Qt.WindowStaysOnTopHint)
        file_dialog.setModal(True)
        file_dialog.setFileMode(QFileDialog.AnyFile)
        file_dialog.setAcceptMode(QFileDialog.AcceptOpen)
        # file_dialog.setDirectory(default_path)
        if file_dialog.exec_():
            selected_file = file_dialog.selectedFiles()[0]
            file_path = QFileInfo(selected_file).path()
            file_name = QFileInfo(selected_file).fileName()
            self.exc_path_edit.setText(selected_file)
            self.file = selected_file

    @pyqtSlot()
    def on_load_data_button_clicked(self):

        self.exc_twidget.setRowCount(0)
        if not self.file:
            return

        head_row = self.header_row_sbox.value()
        data_row = self.data_row_sbox.value()
        data_column = self.data_column_sbox.value()

        wb = openpyxl.load_workbook(os.path.join(os.getcwd(), self.file), read_only=True)
        ws = wb.active
        # print ws
        # print ws[head_row]
        headers = [unicode(item.value) for item in ws[head_row] if item.value is not None]

        # listID = [self.listWidgetID.item(i).text() for i in range(self.listWidgetID.count())]

        data = ws.iter_rows(row_offset=data_row, column_offset=data_column)

        self.exc_twidget.setColumnCount(len(headers))

        self.exc_twidget.setHorizontalHeaderLabels(headers)

        for x, rows in enumerate(data):
            if rows[0].value is not None:
                # if str(rows[0].value) in listID:
                self.exc_twidget.setRowCount(self.exc_twidget.rowCount() + 1)
                for y, cell in enumerate(rows):
                    val = cell.value
                    if val is not None:
                        item = QTableWidgetItem(unicode(val))
                        self.exc_twidget.setItem(self.exc_twidget.rowCount() - 1, y, item)